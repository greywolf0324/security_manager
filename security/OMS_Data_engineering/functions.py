import datetime
import uuid
from .OCR.parse import BUCEE_Parsing, PEPCO_Parsing, PEPCO_Add_Parsing, Walgreens_Parsing, Dollarama_Parsing, Family_Dollar_Parsing, Gabes_Parsing, TEDI_Parsing, Walmart_Parsing, Ollies_Parsing, ORBICO_Parsing, EXCEL_Parsing, CVS_Parsing, GiantTiger_Parsing, HOBBYlobby_Parsing, Lekia_Parsing, Byebye_Parsing, DollarTree_Parsing
from .Original_SalesImport_Generation.matching import PO_Match_BUCEE, PO_Match_PEPCO, PO_Match_PEPCO_Add, PO_Match_Walgreens, PO_Match_Dollarama, PO_Match_Family_Dollar, PO_Match_Gabes, PO_Match_TEDI, PO_Match_Walmart, PO_Match_Ollies, PO_Match_ORBICO, PO_Match_EXCEL, PO_Match_CVS, PO_Match_GiantTiger, PO_Match_HOBBYlobby, PO_Match_Lekia, PO_Match_ByebyeBaby, PO_Match_DollarTree
from .DB_Updater.equal_extractor import Extractor
from .DB_Updater.customers_updater import customer_fields_updater
from .DB_Updater.auto_SKU import AutoDB
from .SalesImport_Generation.Integrator import Integrate_All
from .SalesImport_Generation.SalesImport_updater import SalesImport_Updater
from csv import DictWriter
import json
import pandas as pd
import csv
from openpyxl import load_workbook
import xlsxwriter
import os
from pathlib import Path
from os.path import exists
import re
from decimal import Decimal
from google.oauth2 import service_account
import gspread
from .sheet_reader import frame_converter
from ..models import Customers, Original_SalesImport, Matching_dict, Osalesimport_fields, OMS_Customers, OMS_Payment_term, OMS_AdditionalUOM, OMS_UOM, OMS_Locations, OMS_Inventory_List, Input_paths, Ocustomer_fields, Oinventory_fields, Olocation_fields
from openpyxl import Workbook
from openpyxl.styles import numbers
from ..models import Osalesimport_fields, SalesImport_fields
from collections import OrderedDict
from .util import orderer, field_adder, modelto_dataframe, monday_pagefetcher
from monday import MondayClient

filenames = []

class SalesImport_Generator:
    def __init__(self) -> None:
        # self.mondayoms_fetcher()

        self.customer_name = ""
        self.auto_dic = []
        self.matching_res = []
        self.SKU_list = [item['customer_name'] for item in Customers.objects.filter(sku_bl=True).values('customer_name')]
        self.matching_cols = [f.name for f in Original_SalesImport._meta.get_fields() if f.name != 'id'][3:]

        self.osales_fieldnames = [f.field_name for f in Osalesimport_fields.objects.all()]
        self.sales_fieldnames = [f.field_names for f in SalesImport_fields.objects.all()]
        self.customer_SKU_list = ["Pepco", "Poundland", "Walgreens", "Ollies", "CVS", "Giant Tiger", "Hobby Lobby", "Dollarama"]
        self.header_keys = {
            "PO#": "PO Number", 
            "PO Date": "PO Date", 
            "Dept": "Dept #", 
            "Ship Date": "Ship Dates", 
            "Cancel Date": "Cancel Date", 
            "Payment Terms": "Payment Terms Disc Days Due",
            "PO Total": ""
        } # CustomerName exists
        self.item_keys = {
            "Buyers Catalog or Stock Keeping #": "Buyers Catalog or Stock Keeping #",
            "UPC": "UPC/EAN", 
            "Vendor Style": "Vendor Style", 
            "Retail Price": "Retail Price",  
            "Unit Of Measure": "Unit of Measure", 
            "Number of Pcs per Case Pack": "Pack Size UOM",
            "Quantity Ordered": "Qty Ordered",
            "Total Case Pack Qty": "",
            "Unit Price": "Unit Price", 
            "Pack Size": "", 
            "Number of Pcs per Inner Pack": "Number of Pcs per Inner Pack",
            "Number of Inner Packs": "Number of Inner Packs",  
            "Price Total Amount": ""
        }
        self.input_item_keys = ["Buyers Catalog or Stock Keeping #", "UPC", "Vendor Style", "Retail Price", "Unit Of Measure", "Unit Price", "Quantity Ordered", "Total Case Pack Qty", "Pack Size", "Number of Pcs per Case Pack", "Number of Pcs per Inner Pack", "Number of Inner Packs", "Price Total Amount"]

        self.inventory_matching = modelto_dataframe(OMS_Inventory_List, Oinventory_fields)
        self.stocklocations = modelto_dataframe(OMS_Locations, Olocation_fields)
        self.customers = modelto_dataframe(OMS_Customers, Ocustomer_fields)

    def mondayoms_fetcher(self):
        Customers_boardID = 5790363144
        Inventory_boardID = 5829229370
        StockLocation_boardID = 5751350023

        x = OMS_Customers.objects.all()
        x.delete()

        customers = monday_pagefetcher(Customers_boardID)

        for customer in customers:
            input = OMS_Customers(
                Name = customer["name"],
                Status = customer["column_values"][4]["text"],
                Currency = customer["column_values"][5]["text"],
                PaymentTerm = customer["column_values"][6]["text"],
                TaxRule = customer["column_values"][7]["text"],
                PriceTier = customer["column_values"][8]["text"],
                Discount = customer["column_values"][9]["text"],
                CreditLimit = customer["column_values"][10]["text"],
                Carrier = customer["column_values"][12]["text"],
                SalesRepresentative = customer["column_values"][13]["text"],
                Location = customer["column_values"][14]["text"],
                TaxNumber = customer["column_values"][15]["text"],
                Tags = customer["column_values"][16]["text"],
            )
            input.save()

        x = OMS_Inventory_List.objects.all()
        x.delete()

        inventories = monday_pagefetcher(Inventory_boardID)

        for inventory in inventories:
            input = OMS_Inventory_List(
                ProductCode = inventory["name"],
                Name = inventory["column_values"][0]["text"],
                Type = inventory["column_values"][1]["text"],
                Brand = inventory["column_values"][3]["text"],
                CostingMethod = inventory["column_values"][4]["text"],
                DefaultUnitOfMeasure = inventory["column_values"][5]["text"],
                PurchaseTaxRule = inventory["column_values"][10]["text"],
                SaleTaxRule = inventory["column_values"][11]["text"],
                Status = inventory["column_values"][17]["text"],
                DefaultLocation = inventory["column_values"][18]["text"],
                Length = inventory["column_values"][40]["text"],
                Width = inventory["column_values"][41]["text"],
                Height = inventory["column_values"][42]["text"],
                Weight = inventory["column_values"][43]["text"],
                WeightUnits = inventory["column_values"][44]["text"],


                # Length = inventory["column_values"][6]["text"],
                # Width = inventory["column_values"][7]["text"],
                # Height = inventory["column_values"][8]["text"],
                # Weight = inventory["column_values"][9]["text"],
                # CartonLength = inventory["column_values"][10]["text"],
                # CartonWidth = inventory["column_values"][11]["text"],
                # CartonHeight = inventory["column_values"][12]["text"],
                # CartonInnerQuantity = inventory["column_values"][13]["text"],
                # CartonQuantity = inventory["column_values"][14]["text"],
                # CartonVolume = inventory["column_values"][15]["text"],
                # WeightUnits = inventory["column_values"][16]["text"],
                # DimensionUnits = inventory["column_values"][17]["text"],
                # Barcode = inventory["column_values"][18]["text"],
                # ReorderQuantity = inventory["column_values"][20]["text"],
                # DefaultLocation = inventory["column_values"][21]["text"],
                # LastSuppliedBy = inventory["column_values"][22]["text"],
                # SupplierProductCode = inventory["column_values"][23]["text"],
                # SupplierProductName = inventory["column_values"][24]["text"],
                # SupplierFixedPrice = inventory["column_values"][25]["text"],
                # PriceTier1 = inventory["column_values"][26]["text"],
                # PriceTier2 = inventory["column_values"][27]["text"],
                # PriceTier3 = inventory["column_values"][28]["text"],
                # PriceTier4 = inventory["column_values"][29]["text"],
                # PriceTier5 = inventory["column_values"][30]["text"],
                # PriceTier6 = inventory["column_values"][31]["text"],
                # PriceTier7 = inventory["column_values"][32]["text"],
                # PriceTier8 = inventory["column_values"][33]["text"],
                # PriceTier9 = inventory["column_values"][34]["text"],
                # PriceTier10 = inventory["column_values"][35]["text"],
                # AssemblyBOM = inventory["column_values"][36]["text"],
                # AutoAssemble = inventory["column_values"][37]["text"],
                # AutoDisassemble = inventory["column_values"][38]["text"],
                # DropShip = inventory["column_values"][39]["text"],
                # DropShipSupplier = inventory["column_values"][40]["text"],
                # AverageCost = inventory["column_values"][41]["text"],
                
                # InventoryAccount = inventory["column_values"][43]["text"],
                # RevenueAccount = inventory["column_values"][44]["text"],
                # ExpenseAccount = inventory["column_values"][45]["text"],
                # COGSAccount = inventory["column_values"][46]["text"],
                # ProductAttributeSet = inventory["column_values"][47]["text"],
            )
        
            input.save()

        x = OMS_Locations.objects.all()
        x.delete()
        
        locations = monday_pagefetcher(StockLocation_boardID)

        for location in locations:
            input = OMS_Locations(
                Location = location["name"]
            )
            input.save()

    def str_converter(self, input):
        temp = []
        for item in list(input):
            temp.append(str(item))

        return temp
    
    def auto_matching_DB_viewer(self, customer_name):
        # self.inventory_matching = modelto_dataframe(OMS_Inventory_List.objects.all())
        if exists(Path(__file__).resolve().parent / f"config/AutoFill_DB/{customer_name}.csv"):
            
            auto_df = pd.read_csv(Path(__file__).resolve().parent / f"config/AutoFill_DB/{customer_name}.csv", index_col = False)
            temp = []
            for item in auto_df["PO"]:
                if type(item) == float:
                    temp.append(int(item))
                else:
                    temp.append(item)

            UOM_options = ["Case", "Each"]
            location_options = list(self.stocklocations["Locations"])
            vendor_options = {}
            if customer_name in self.SKU_list:
                for sku in auto_df["PO"]:
                    temp = []
                    for item in self.inventory_matching["ProductCode"]:
                        if str(sku) in item:
                            temp.append(item)
                    vendor_options.update({
                        str(sku): temp
                    })
            else:
                for sku in auto_df["PO"]:
                    vendor_options.update({
                        str(sku): list(self.inventory_matching["ProductCode"])
                    })
            print("________________________")
            target = self.str_converter(auto_df["Vendor Style from OMS_equal"])
            print("========================")

            return {"PO": list(auto_df["PO"]), "UOM": list(auto_df["Unit of Measure"]), "LOCATION": list(auto_df["StockLocation"]), "TARGET": target, "UOM_options": UOM_options, "location_options": location_options, "vendor_options": vendor_options}
        
        else: 
            return "non-exists"
        
    def auto_matching_DB_changer(self, customer_name, db):
        print("updating vendor style matching DB...")
        pd.DataFrame(db[0]).to_csv(Path(__file__).resolve().parent / f"config/AutoFill_DB/{customer_name}.csv", index = False)

    def uploadFile(self, data):
        self.paths = []            
        global filenames
        uuid_code = str(uuid.uuid4())
        for file in data:
            obj = Input_paths()
            current_datetime = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            random_string = str(uuid.uuid4().hex)
            filename = f'{current_datetime}_{random_string}'
            filenames.append(filename)
            obj.file_name = str(filename)
            extension = file.name.split(".")[-1]
            path = Path(__file__).resolve().parent.parent.parent / f'process/inputs/{filename}.{extension}'
            
            self.paths.append(path)
            obj.path = str(path)
            obj.created = uuid_code
            obj.save()

            with open(path, 'wb') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
    
    def parseUpload(self, data, customer_name = None, currency = None):
        print("==============================================================================================================")
        print("CustomerFields Updating...")

        customer_fields_updater()

        self.customer_name = customer_name
        self.uploadFile(data)
        paths = self.paths

        print("==============================================================================================================")
        print("On PDF parsing...")
        parser = eval(Matching_dict.objects.filter(customer_name = customer_name)[0].parser)(customer_name)
        PO_res = parser.PO_parser(paths, currency)
        print(PO_res[0])

        print("==============================================================================================================")
        print("On Match Operating...")
        matcher = eval(Matching_dict.objects.filter(customer_name = customer_name)[0].matcher)()
        matching_res = matcher.match_final(PO_res)
        self.matching_res = matching_res
        # print(matching_res[0])
        uuid_code = str(uuid.uuid4())
        matching_res = orderer(matching_res, self.osales_fieldnames, customer_name)
        
        for k, _ in enumerate(matching_res):
            length = len(matching_res[k][list(matching_res[k].keys())[0]])
            for i in range(length):
                if i == 0:
                    obj = Original_SalesImport.objects.create(start_line = True)
                else:
                    obj = Original_SalesImport.objects.create(start_line = False)

                obj.created = uuid_code
                for key, vkey in zip(self.matching_cols, matching_res[k].keys()):
                    if key in ["PO_Number", "Release_Number", "Retailers_PO", "Buyers_Catalog_or_Stock_Keeping", "UPC_EAN", "Ship_To_Location", "PO_Line", "Qty_Ordered", "Vendor_Style", "Number_of_Inner_Packs", "Number_of_Pcs_per_Inner_Pack", "Qty_per_Store", "Vendor", "Ship_to_Zip", "Bill_To_Zip", "Buying_Party_Zip", "Mark_for_Postal"]:
                        try:
                            matching_res[k][vkey][i] = int(float(matching_res[k][vkey][i]))
                        except:
                            if matching_res[k][vkey][i] == '':
                                matching_res[k][vkey][i] = None
                            else:
                                pass
                    elif key in ["Unit_Price", "Retail_Price", "PO_Total_Amount", "PO_Total_Weight"]:
                        try:
                            matching_res[k][vkey][i] = float(matching_res[k][vkey][i])
                        except:
                            matching_res[k][vkey][i] = None

                    elif key in ["PO_Date","Requested_Delivery_Date","Delivery_Dates","Ship_Dates","Cancel_Date"]:
                        if customer_name in ["Buc-ee's", "Big Lots Stores", "CVS", "Five Below", "Fred Meyer", "Meijers", "MICHAELS", "Tar Heel Trading", "TARGET", "Walgreens", "Walmart US", "Gabe's", "Hobby Lobby", "Ollies", "Walmart", "Dollarama", "Family Dollar", "Dollar Tree Stores"]:
                            try:
                                temp = matching_res[k][vkey][i].split("/")
                                matching_res[k][vkey][i] = "-".join([temp[i] for i in [2, 0, 1]])
                            except:
                                matching_res[k][vkey][i] = None
                        
                        elif customer_name in ["Pepco", "Poundland", "buy buy BABY", "TEDI"]:
                            try:
                                temp = matching_res[k][vkey][i].split("/")
                                matching_res[k][vkey][i] = "-".join([temp[i] for i in [2, 1, 0]])
                            except:
                                matching_res[k][vkey][i] = None
                                
                    else:
                        if matching_res[k][vkey][i] == "":
                            matching_res[k][vkey][i] = None
                        
                        else:
                            pass

                    setattr(obj, key, matching_res[k][vkey][i])
                
                obj.save()
        
        # # Extract equal OMS
        print("==============================================================================================================")
        print("tracking equal things...")
        extract = Extractor()
        OMS_equal = extract.extractor(matching_res, customer_name)
        self.auto_dic = AutoDB().DB_tester(customer_name, matching_res)

        print("==============================================================================================================")
        print("Displaying...")
        if (self.customer_name == "Pepco" or self.customer_name == "Poundland") and matching_res[0]["Currency"][0] == "CNY":
                self.customer_name = self.customer_name + " - RMB"
        else:
            if self.customer_name == "Pepco":
                self.customer_name = self.customer_name + " - " + matching_res[0]["Currency"][0]
            
            elif self.customer_name == "Walmart":
                if matching_res[0]["Currency"][0] == "US Dollar":
                    self.customer_name = self.customer_name + " US"
        
        return [matching_res, OMS_equal, self.auto_dic, list(self.stocklocations["Locations"]), self.customer_name]

    def res_viewer(self, data, matching_res, customer_name = None, term = None):
        objs = Input_paths.objects.filter(created = Input_paths.objects.all()[len(Input_paths.objects.all()) - 1].created)
        filename = objs[0].file_name

        #Fields being filled from selected Vendor Style: Pack Size UOM, Number of Inner Packs, Number of Pcs per Inner Pack
        for invoice in matching_res:
            for i in range(len(invoice[list(invoice.keys())[0]])):
                if i != 0:
                    self.inventory_matching["ProductCode"] = self.inventory_matching["ProductCode"].astype('str')
                    temp = self.inventory_matching[self.inventory_matching["ProductCode"] == invoice["Vendor Style"][i]]["DefaultUnitOfMeasure"]

                    if temp.values[0] == "Unit":
                        invoice["Pack Size UOM"][i] = 1
                        invoice["Number of Pcs per Inner Pack"][i] = 1
                        invoice["Number of Inner Packs"][i] = 1
                    
                    else:
                        nums = re.sub(r"[a-zA-Z]", "", temp.values[0]).replace(" ", "").split(",")
                        try:
                            invoice["Pack Size UOM"][i] = nums[0]
                            invoice["Number of Pcs per Inner Pack"][i] = nums[1]
                            invoice["Number of Inner Packs"][i] = int(int(float(invoice["Pack Size UOM"][i])) / int(float(invoice["Number of Pcs per Inner Pack"][i])))
                        except:
                            invoice["Pack Size UOM"][i] = nums[0]
                            invoice["Number of Pcs per Inner Pack"][i] = 1
                            invoice["Number of Inner Packs"][i] = int(int(float(invoice["Pack Size UOM"][i])) / int(float(invoice["Number of Pcs per Inner Pack"][i])))

        print("==============================================================================================================")
        print("Creating View...\n")
        header_details = []
        item_details = []
        temp_item_details = []
        for i in range(len(matching_res)):
            header_details.append({})
            for key in self.header_keys:
                header_details[i].update(
                    {
                        key: []
                    }
                )
            temp_item_details.append({})
            for key in self.item_keys:
                temp_item_details[i].update(
                    {
                        key: []
                    }
                )
        PO_total = []

        for i, invoice in enumerate(matching_res):
            PO_total.append(0)
            for key in self.item_keys:
                if key == "Price Total Amount":
                    for j in range(len(invoice[self.item_keys["UPC"]]) - 1):
                        try:
                            PO_total[i] += float(Decimal(str(float(invoice["Unit Price"][j + 1]))) * int(float(invoice["Qty Ordered"][j + 1])))
                        except:
                            PO_total[i] += float(Decimal(str(float(invoice["Unit Price"][1]))) * int(float(invoice["Qty Ordered"][j + 1])))
                        temp_item_details[i][key].append(float(Decimal(str(float(invoice["Unit Price"][1]))) * int(float(invoice["Qty Ordered"][j + 1]))))
                elif key == "Pack Size":
                    for j in range(len(invoice[self.item_keys["UPC"]]) - 1):
                        temp_item_details[i][key].append(1)

                elif key == "Quantity Ordered":
                    for item in invoice[self.item_keys[key]][1:]:
                        temp_item_details[i][key].append(int(float(item)))

                elif key == "Number of Pcs per Case Pack":
                    for item in invoice[self.item_keys[key]][1:]:
                        temp_item_details[i][key].append(item)

                elif key == "Total Case Pack Qty":
                    if invoice[self.item_keys["Unit Of Measure"]][1] == "Case":
                        for j in range(len(invoice[self.item_keys["UPC"]]) - 1):
                            temp_item_details[i][key].append(temp_item_details[i]["Quantity Ordered"][j])
                    else:
                        for j in range(len(invoice[self.item_keys["UPC"]]) - 1):
                            temp_item_details[i][key].append(temp_item_details[i]["Quantity Ordered"][j] / int(temp_item_details[i]["Number of Pcs per Case Pack"][j]))

                elif key == "Unit Of Measure":
                    if invoice[self.item_keys[key]][1] == "Case":
                        for item in invoice[self.item_keys[key]][1:]:
                            temp_item_details[i][key].append(item)

                    else:
                        for item in invoice[self.item_keys[key]][1:]:
                            temp_item_details[i][key].append("Each")

                else:
                    for item in invoice[self.item_keys[key]][1:]:
                        temp_item_details[i][key].append(item)
        
        for i, invoice in enumerate(matching_res):
            for key in self.header_keys:
                if key == "Payment Terms":
                   header_details[i][key].append(term[0])
                
                elif key == "PO Total":
                    header_details[i][key].append(PO_total[i])

                else: 
                    header_details[i][key].append(invoice[self.header_keys[key]][0])

        for i in range(len(matching_res)):
            item_details.append({})
            for key in self.input_item_keys:
                item_details[i].update(
                    {
                        key: temp_item_details[i][key]
                    }
                )
        res = [customer_name, header_details, item_details]
        with open(Path(__file__).resolve().parent.parent.parent / f'process/views/{filename}.json', 'w') as f:
            json.dump(res, f)
        
        return [customer_name, header_details, item_details]
    
    def processFile(self, data, matching_res, extra, customer_name, terms):
        # f = open(Path(__file__).resolve().parent / "config/django-connection-1008-5f931d8f4038.json")
        # google_json = json.load(f)
        # credentials = service_account.Credentials.from_service_account_info(google_json)
        # scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
        # creds_with_scope = credentials.with_scopes(scope)
        # client = gspread.authorize(creds_with_scope)
        # self.spreadsheet = client.open_by_url("https://docs.google.com/spreadsheets/d/1CDnIivm8hatRQjG7nvbMxG-AuP19T-W2bNAhbFwEuE0")
        obj = Input_paths.objects.filter(created = Input_paths.objects.all()[len(Input_paths.objects.all()) - 1].created)
        path = obj[0].path
        filename = obj[0].file_name
        sku_match = {}

        if customer_name in self.SKU_list:
            for dic, dic_self in zip(self.matching_res, matching_res):
                for key, uom, location, target in zip(list(dic["Vendor Style"])[1:], list(dic_self["Unit of Measure"])[1:], list(dic_self["StockLocation"])[1:], list(dic_self["Vendor Style"])[1:]):
                    sku_match.update(
                        {
                            str(key): [uom, location, target]
                        }
                    )
        
        else:
            for dic, dic_self in zip(self.matching_res, matching_res):
                for key, uom, location, target in zip(list(dic["Buyers Catalog or Stock Keeping #"])[1:], list(dic_self["Unit of Measure"])[1:], list(dic_self["StockLocation"])[1:], list(dic_self["Vendor Style"])[1:]):
                    sku_match.update(
                        {
                            str(key): [uom, location, target]
                        }
                    )
            
        print("==============================================================================================================")
        print("updating Auto matching DB ...")
        AutoDB().auto_DB_updater(sku_match, customer_name)

        print("==============================================================================================================")
        print("Integrating...")
        integrator = Integrate_All(customer_name=customer_name)
        sales_import = integrator.Integrate_final(matching_res, customer_name, terms)
        # print(sales_import[0])
        print("==============================================================================================================")
        print("Updating SalesImport...")
        updater = SalesImport_Updater()
        sales_import = updater.updater(sales_import)

        print("==============================================================================================================")
        print("Just a second, writing...")
        f = open(Path(__file__).resolve().parent / "config/fieldnames_SalesImport.json")
        o_field_names = json.load(f)
        wb = Workbook()
        ws = wb.active

        # Example decimal numbers
        # number = [0.500, 1.000, 1.500]
        sales_import = orderer(field_adder(sales_import, self.sales_fieldnames), self.sales_fieldnames, customer_name)
        # print(sales_import[0])
        for i, key in enumerate(self.sales_fieldnames):
            cell = ws.cell(row = 1, column = i + 1, value = key)

        counter = 1
        for i, order in enumerate(sales_import):
            len_order = len(order[list(order.keys())[0]])
            for j in range(len_order):
                counter += 1
                for k, key in enumerate(order):
                    if key == "Discount":
                        try:
                            cell = ws.cell(row = counter, column = k + 1, value = float(order[key][j]))
                            cell.number_format = numbers.FORMAT_NUMBER_00
                        except:
                            pass
                    elif key == "Total*":
                        cell = ws.cell(row = counter, column = k + 1, value = order[key][j])
                        cell.number_format = numbers.FORMAT_NUMBER_00
                    else:
                        cell = ws.cell(row = counter, column = k + 1, value = order[key][j])
                    # cell = ws.cell(row = counter, column = k + 1, value = order[key][j])
                    # if key in ["Discount", "Total*"]:
                    #     if key == "Discount":
                    #         try:

                    #         print(order[key][j], type(order[key][j]))
                    #     cell.number_format = numbers.FORMAT_NUMBER_00
                    
                
            # for j, key in enumerate(order):
            #     print(order[key])
                        
        # # Write numbers to the sheet with trailing zeros preserved
        # for i, numbe in enumerate(number):
        #     cell = ws.cell(row=i+1, column=1, value=numbe)
        #     cell.number_format = numbers.FORMAT_NUMBER_00

        # Save the workbook
        wb.save(Path(__file__).resolve().parent.parent.parent / f'process/outputs/{filename}.csv')
        # if os.path.isfile("SalesImport.xlsx"):
        #     os.remove("SalesImport.xlsx")
        # book = xlsxwriter.Workbook("SalesImport.xlsx")
        # sheet = book.add_worksheet("cont_excel")

        # for idx, header in enumerate(o_field_names):
        #     sheet.write(0, idx, header)
        # book.close()
        # book = load_workbook("SalesImport.xlsx")
        # sheet = book.get_sheet_by_name("cont_excel")
        # for _, dic in enumerate(sales_import):
        #     keys = list(dic.keys())
        #     for i in range(len(dic[keys[0]])):
        #         temp = []
        #         for key in o_field_names:
        #             if key in keys:
        #                 temp.append(dic[key][i])
        #             else:
        #                 temp.append("")
        #         sheet.append(temp)
        # output = Path(__file__).resolve().parent.parent.parent / f'process/outputs/{filename}.xlsx'
        
        # book.save(filename = output)
        # df = pd.read_excel(output)
        # df.to_excel(Path(__file__).resolve().parent.parent.parent / f'process/outputs/{filename}.xlsx', index=False, float_format='%.2f')

        output = Path(__file__).resolve().parent.parent.parent / f'process/outputs/{filename}.csv'

        return [path, output]
    
    def live_save(self, matching_res, customername, terms):
        worksheet = self.spreadsheet.get_worksheet(6)
        leng = int(self.spreadsheet.get_worksheet(7).get_values()[0][0])
        f = open(Path(__file__).resolve().parent / "config/fieldnames_SalesImport.json")
        #Getting Invoice Numbers
        dt = worksheet.get_all_records()
        Inumbers = []
        for invoice in dt:
            Inumbers.append(invoice["InvoiceNumber*"])
        integrator = Integrate_All(customer_name=customername)
        sales_import = integrator.Integrate_final(matching_res, customername, terms)
        print("==============================================================================================================")
        print("Updating SalesImport...")
        updater = SalesImport_Updater()
        sales_import = updater.updater(sales_import)
        field_names = json.load(f)
        
        Ignore_invoice_nums = []
        Ignore_invoices = []
        res = []
        for dic in sales_import:
            if dic["InvoiceNumber*"][0] in Inumbers:
                Ignore_invoice_nums.append(dic["InvoiceNumber*"][0])
                Ignore_invoices.append(dic)
                continue
            length = len(dic[list(dic.keys())[0]])
            keys = list(dic.keys())
            for i in range(length):
                temp = []
                for key in field_names:
                    if key in keys:
                        temp.append(str(dic[key][i]))
                    else:
                        temp.append("")
                
                res.append(temp)
        worksheet.update(f'{leng + 1}:{leng + len(res)}', res)
        write_sheet = self.spreadsheet.get_worksheet(7)
        write_sheet.update(f'{1}:{2}', [[leng + len(res)]])
        print("success!!!!!!!!!!!!!!!")
        return [Ignore_invoices, sales_import, Ignore_invoice_nums]
    
    def live_addition(self, Ignore_invoices, sales_import, invoice_nums):
        print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        worksheet = self.spreadsheet.get_worksheet(6)
        leng = int(self.spreadsheet.get_worksheet(7).get_values()[0][0])
        f = open(Path(__file__).resolve().parent / "config/fieldnames_SalesImport.json")
        updater = SalesImport_Updater()
        sales_import = updater.updater(sales_import)
        field_names = json.load(f)
        dt = worksheet.get_all_records()
        Inumbers = []
        for i, invoice in enumerate(dt):
            if invoice["InvoiceNumber*"] in invoice_nums:
                Inumbers.append(i)
        res = []
        for dic in Ignore_invoices:
            length = len(dic[list(dic.keys())[0]])
            keys = list(dic.keys())
            for i in range(length):
                temp = []
                for key in field_names:
                    if key in keys:
                        temp.append(str(dic[key][i]))
                    else:
                        temp.append("")
                
                res.append(temp)
        for i, liveline_num in enumerate(Inumbers):
            cells = worksheet.range(f"A{liveline_num + 2}:BG{liveline_num + 2}")
            for j, e in enumerate(cells):
                e.value = res[i][j]
            worksheet.update_cells(cells)
        print("Success Live Addition!!!")